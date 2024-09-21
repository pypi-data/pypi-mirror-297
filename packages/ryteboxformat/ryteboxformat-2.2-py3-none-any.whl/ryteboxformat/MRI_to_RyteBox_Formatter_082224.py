#! /usr/bin/env py -3
from audioop import error
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
import sys
import multiprocessing
import os
import logging
 
# Main Function 
def main(config_file_path):

	# Get the directory of the current script
	script_dir = os.path.dirname(os.path.abspath(__file__))
	#logging.debug('got here 1')
	# Retrieve arguments passed from the calling script
	#config_file_path = sys.argv[0]
	logging.debug('FM Code: config_file_path from Arg passed in: ' + str(config_file_path))
	# Load the config file and extract specific formatting config settings
	config_wb = openpyxl.load_workbook(config_file_path / 'user_data.xlsx')
	config_ws = config_wb.active


	if getattr(sys, 'frozen', False):
		script_dir = os.path.dirname(sys.executable)
	else:
		script_dir = os.path.dirname(os.path.abspath(__file__))
		
	# Change the working directory to the script's directory
	logging.debug('FM: WORKING DIR: ' + script_dir)
	os.chdir(script_dir)
		
	# Now we can use relative paths
	mappingFile = os.path.join(script_dir, 'Config Master 081424.xlsx')
	logging.debug('FM: NEW Mapping File PATH: ' + mappingFile)

	# This section should be passed in as arguments
	sourceFile = config_ws.cell(row=2, column=1).value
	sourceFileTab = config_ws.cell(row=2, column=4).value
	sourceFileHeaderRow = config_ws.cell(row=2, column=5).value
	sourceFileMaxRow = config_ws.cell(row=2, column=6).value # the row the total line is on minus on, or said another way, the last row of data to process
	# mappingFile = '.\Config Master 081424.xlsx' #this assumes this file is in the same dir as this py script
	mappingFileTab = config_ws.cell(row=2, column=7).value + ' Config'
	mappingFileHeaderRow = 1
	targetFile = os.path.join(config_file_path, (mappingFileTab.upper() + '_Output_' + str(datetime.now().strftime("%Y%m%d%H%M%S")))) + '.xlsx'
	#logging.debug(targetFile)

	
	# Load the source and mapping files
	source_wb = openpyxl.load_workbook(sourceFile) # + '.xlsx' removed
	source_ws = source_wb[sourceFileTab]
	
	mapping_wb = openpyxl.load_workbook(mappingFile)
	mapping_ws = mapping_wb[mappingFileTab]
	
	# Create a new workbook for the target file
	target_wb = openpyxl.Workbook()
	# target_wb.save(targetFile)
	
	# Open the target file spreadsheet
	# target_wb = openpyxl.Workbook(targetFile)
	target_ws = target_wb.active
	
	# Write the RyteBox Format headers
	target_ws.append(['Payee/Member Id', 'Payee/Member Name', 'Payee/Member IPI', 'ISWC', 'ISRC', 'Identifier', 'Third party ID', 'Title', 'Contributors', 'Source', 'Source Details', 'Amount', 'Units', 'Start Date', 'End Date', 'Usage Territory Name', 'Usage Territory Code', 'Usage Type', 'Usage Type Code', 'Recording Title', 'Catalog #', 'Album Title', 'Artist', 'Show Name', 'Episode Name', 'Notes', 'Share % Received'])
	
	# Define styles
	date_style = NamedStyle(name="date_style", number_format="MM/DD/YYYY")
	text_style = NamedStyle(name="text_style", number_format="Text")
	
	# Read the mapping file and apply the mappings
	logging.debug('FM: Beginning mapping...')
	for row in mapping_ws.iter_rows(min_row=2, values_only=True):
		i=1
		source_col, target_col, target_format, target_value = row
		#logging.debug(source_col)
		if source_col is not None:
			#logging.debug(source_col)
			if source_col != 'Constant': source_col_idx = openpyxl.utils.column_index_from_string(source_col)
			#logging.debug('source col idx : ' + str(source_col_idx))
			target_col_idx = openpyxl.utils.column_index_from_string(target_col)
			#logging.debug('target col idx : ' + str(target_col_idx))
	
		# Apply the mappings to the target file
		if source_col is not None:
			for i, cell in enumerate(source_ws.iter_rows(min_col=source_col_idx, max_col=source_col_idx, min_row=int(sourceFileHeaderRow)+1, max_row=int(sourceFileMaxRow), values_only=True), start=2):
				# logging.debug(str(i) + " - "+ str(cell))
				#if cell is not None:
				
				if source_col == 'Constant':
					target_cell = target_ws.cell(row=i, column=target_col_idx, value = target_value)
				else:
					target_cell = target_ws.cell(row=i, column=target_col_idx, value = formattedCell(target_format, cell[0])) # value=source_ws.cell(row=i, column=source_col_idx)
		
				# Apply formatting	
				if target_format == "mm/dd/yyyy":
					target_cell.style = date_style
				elif target_format == "text":
					target_cell.style = text_style		
				# target_wb.save(targetFile)
	
	# Save the target file
	logging.debug('FM: Mapping try completed')
	target_wb.save(f'{targetFile}')
	logging.debug(f'FM: Target XLSX file created: {targetFile}')
	logging.debug('FM: CSV created') if createCSV(targetFile) else logging.error('FM: CSV creation failed')
	logging.debug('FM: -done-')
	
def formattedCell(target_format, cellValue):
	# Determine which format converter function to call
	if target_format == 'convertDate_MMM_YYYY_to_RB_Date': 
		return convertDate_MMM_YYYY_to_RB_Date(cellValue)
	elif target_format == 'convertValue_Strip%_MakeFloat': 
		return convertValue_StripPercemt_MakeFloat(cellValue)
	elif target_format == 'convertDate_YYYYQQ_to_RB_StartDate': 
		return convertDate_YYYYQQ_to_RB_StartDate(cellValue)
	elif target_format == 'convertDate_YYYYQQ_to_RB_EndDate': 
		return convertDate_YYYYQQ_to_RB_EndDate(cellValue)
	elif target_format == 'shortDate_to_RB_Date':
		return convertDate_m_d_yy_to_RB_Date(cellValue)
	elif target_format == 'extractDate_Income_Period_to_RB_Start_Date':
		return extractDate_Income_Period_to_RB_Start_Date(cellValue)	
	elif target_format == 'extractDate_Income_Period_to_RB_End_Date':
		return extractDate_Income_Period_to_RB_End_Date(cellValue)	
	else: return cellValue

def convertDate_MMM_YYYY_to_RB_Date(date_str):
	# Convert string to datetime object
	if date_str is not None:
		date_obj = datetime.strptime(date_str, '%b %Y')
		# Format datetime object to desired format
		formatted_date = date_obj.strftime('%m/01/%y')
		return formatted_date
	else: 
		return None
		
def convertDate_YYYYQQ_to_RB_StartDate(date_str):
	# Parse string to datetime pieces, then convert to date object
	if date_str is not None:
		parts = date_str.upper().split('Q')
		dt = datetime(int(parts[0]), int(parts[1])*3-2, 1)
		formatted_date = dt.strftime('%m/01/%y')
		return formatted_date
	else: 
		return None

def convertDate_YYYYQQ_to_RB_EndDate(date_str):
	# Parse string to datetime pieces, then convert to date object
	if date_str is not None:
		parts = date_str.upper().split('Q')
		dt = datetime(int(parts[0]), int(parts[1])*3, 1)
		res = dt + relativedelta(day=31)
		formatted_date = res.strftime('%m/%d/%y')
		return formatted_date
	else: 
		return None
		
def convertDate_m_d_yy_to_RB_Date(date_str):
	# Convert to YYYY/MM/DD format
	if date_str is not None:
		dt = datetime.strptime(date_str, "%Y-%m-%d")
		formatted_date = dt.strftime("%m/%d/%y")
		return formatted_date
	else: 
		return None
		
def extractDate_Income_Period_to_RB_Start_Date(date_str):
# extract the Start date from a period dates string eg., 202301202303.0
	if date_str is not None:
		start_date_str = date_str[:6] # get the first 6 chars
		start_date = datetime.strptime(start_date_str, '%Y%m')	# convert it to datetime objects
		formatted_start_date = start_date.strftime('%m/%d/%Y')	#format the dates into RB date format and return
		return formatted_start_date
	else:
		return None

def extractDate_Income_Period_to_RB_End_Date(date_str):
# extract the End date from a period dates string eg., 202301202303.0
	if date_str is not None:
		end_date_str = date_str[-8:-2] # get the first 5 chars
		end_date = datetime.strptime(end_date_str, '%Y%m')	# convert it to datetime objects
		formatted_end_date = end_date.strftime('%m/%d/%Y')	#format the dates into RB date format and return
		return formatted_end_date
	else:
		return None
		
def convertValue_StripPercemt_MakeFloat(percentVal):
	# Convert percentage value to a float value
	if percentVal is not None:
		share_value = percentVal * 100
		return share_value
	else: 
		return None

def createCSV(strXlSXFile):
	try:
		# Read the Excel file, specifying dtype to ensure text fields with numeric values are read as strings
		df = pd.read_excel(strXlSXFile, dtype=str)
		parts = strXlSXFile.upper().split('XLSX')
		# Export to CSV with float_format to prevent scientific notation
		df.to_csv(str(parts[0]) + '.csv', index=False, float_format='%.10f')
		return True
	except Exception as error:
		logging.error('FM: file: ' + strXlSXFile + ' failed to convert to CSV with error: ' + str(error))
	return False
	
if __name__ == "__main__":
    main()
