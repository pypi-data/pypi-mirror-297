import wx
import pandas as pd
import multiprocessing
import MRI_to_RyteBox_Formatter_082224
import os
import sys
import subprocess
from xls2xlsx import XLS2XLSX
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl import load_workbook
import logging
import io
from pathlib import Path

class SpreadsheetApp(wx.Frame):
	def __init__(self, parent, title):
		super(SpreadsheetApp, self).__init__(parent, title=title, size=(1000, 200))
		self.Bind(wx.EVT_CLOSE, self.OnClose) # added to address the issue of WX firing events when the user attempts to close the form/exit the app
		
		#logging.basicConfig(filename='app.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
		
		logging.debug('UI: UI started...')
		## Create Vertical Box
		self.panel = wx.Panel(self)
		self.sizer = wx.BoxSizer(wx.VERTICAL)
		
		self.file_picker = wx.FilePickerCtrl(self.panel, message="Select a spreadsheet", wildcard="Excel files (*.xlsx;*.xls)|*.xlsx;*.xls|CSV files (*.csv)|*.csv|TSV files (*.tsv)|*.tsv")
		#self.file_picker = wx.FilePickerCtrl(self.panel, message="Select a spreadsheet", wildcard="Excel files (*.xlsx;*.xls)|*.xlsx;*.xls")
		self.sizer.Add(self.file_picker, 0, wx.ALL | wx.EXPAND, 5)		
		
		## Create Horizontal Box
		self.hbox1 = wx.BoxSizer(wx.HORIZONTAL)
		
		# Tuple of radio button labels
		radio_labels = ("TikTok", "Sybus", "Peloton", "Amazon Music","Amazon Prime","Amazon Prime Perf","MLC","Exploration")
		
		self.MRI_Choice_label = wx.StaticText(self.panel, label="MRI Income Statement:")
		
		# Create a radio button group
		self.radio_buttons = []
		self.hbox1.Add(self.MRI_Choice_label, flag=wx.ALIGN_CENTER_VERTICAL|wx.ALL, border=5) # Label
		for i, label in enumerate(radio_labels):
			style = wx.RB_GROUP if i == 0 else 0  # First radio button starts the group
			radio_button = wx.RadioButton(self.panel, label=label, style=style)
			radio_button.Disable()
			radio_button.Bind(wx.EVT_RADIOBUTTON, self.on_radio_button)
			self.radio_buttons.append(radio_button)

			self.hbox1.Add(radio_button, flag=wx.EXPAND|wx.ALL, border=5)
		
		
		## Create Horizontal Box
		self.hbox = wx.BoxSizer(wx.HORIZONTAL)
		
		self.tab_choice_label = wx.StaticText(self.panel, label="Select Tab:")
		self.tab_choice = wx.Choice(self.panel, size=(200, -1))
		self.row_count_label = wx.StaticText(self.panel, label="Last Row Number:") # Last row of usable data (e.g. no total row included in count)
		#self.row_count_text = wx.TextCtrl(self.panel, value="0", pos=wx.DefaultPosition, size=(150, -1), validator=wx.DefaultValidator)
		self.row_count_text = wx.TextCtrl(self.panel, value="0", size=(125, -1))
		self.row_count_text.SetToolTip(wx.ToolTip('Enter the number of the last row of data on the tab excluding any non-data rows (e.g. totals) ')) # this doesnt work, not sure why
		
		# Column headings row. Assumes next row is the start of working data to ingest
		self.header_label = wx.StaticText(self.panel, label="Header Row:")
		self.header_row = wx.TextCtrl(self.panel, value="0", size=(125, -1))

		self.hbox.Add(self.tab_choice_label, flag=wx.ALIGN_CENTER_VERTICAL|wx.ALL, border=5)
		self.hbox.Add(self.tab_choice, flag=wx.ALIGN_CENTER_VERTICAL|wx.ALL, border=5)
		self.hbox.Add(self.header_label, flag=wx.ALIGN_CENTER_VERTICAL|wx.ALL, border=5)
		self.hbox.Add(self.header_row, 0, wx.ALL | wx.Left, 5)
		self.hbox.Add(self.row_count_label, flag=wx.ALIGN_CENTER_VERTICAL|wx.ALL, border=5)
		self.hbox.Add(self.row_count_text, 0, wx.ALL | wx.Left, 5)

		self.sizer.Add(self.hbox1, flag=wx.EXPAND|wx.ALL, border=5)
		self.sizer.Add(self.hbox, flag=wx.EXPAND|wx.ALL, border=5)
		
		self.save_button = wx.Button(self.panel, label="Save Details")
		self.sizer.Add(self.save_button, 0, wx.ALL | wx.CENTER, 5)
		
		self.panel.SetSizer(self.sizer)
		
		#self.load_button.Bind(wx.EVT_BUTTON, self.on_load)
		self.file_picker.Bind(wx.EVT_FILEPICKER_CHANGED, self.on_load)
		#self.radio_button.Bind(wx.EVT_CHECKBOX, self.on_checkbox)
		self.tab_choice.Bind(wx.EVT_CHOICE, self.on_tab_select)
		self.save_button.Bind(wx.EVT_BUTTON, self.on_save)
		
		self.spreadsheet = None
		self.file_path = ""
		self.file_name = ""
		self.file_dir_only = ""
		self.selected_tab = ""
		self.radiolabel = ""
		
		self.Show()
	
	def on_load(self, event):
		self.file_path = self.file_picker.GetPath()
		if self.file_path:  # if a file has been selected review the file to see if xls or xlsx
				
			# convert the source file from .xls to .xlsx
			_, file_extension = os.path.splitext(self.file_path)
			
			if file_extension.upper() == '.XLS':
				logging.debug('UI: Converting the income statement selected from an .XLS file to an XLSX file: ' + os.path.basename(self.file_path))
				rtn = wx.MessageBox('The income statement selected is an .XLS file :\n' + os.path.basename(self.file_path) + '\n and is being copied into a separate .XLSX formatted file\n', 'NOTICE: Upgrade "XLS" Spreadsheet Format', wx.OK | wx.ICON_INFORMATION)
				xlsx_path=self.file_path.replace('.xls', '.xlsx')
				try:
					x2x = XLS2XLSX(self.file_path)
					x2x.to_xlsx(xlsx_path)
					self.file_picker.SetPath(xlsx_path)
					self.file_path = xlsx_path
					logging.debug('UI: XLS conversion succeeded. New file is: ' + os.path.basename(self.file_path))
					wx.MessageBox(f"XLS conversion to XLSX Completed\n File saved as {os.path.basename(self.file_path)}", "SUCCESS XLSX Upgrade Complete", wx.OK | wx.ICON_INFORMATION)
				except Exception as e:
					error.debug(f'Failed to convert and /or save the XLS file as an XLSX: {e}')
					wx.MessageBox(f"Failed to save the XLSX file: {e}", "Error", wx.OK | wx.ICON_ERROR)
					
			if file_extension.upper() in ['.CSV', '.TSV']:
				logging.debug(f'UI: Converting the income statement selected from an {file_extension.upper()} file to an XLSX file: {os.path.basename(self.file_path)}')
				rtn = wx.MessageBox(f'The income statement selected is an {file_extension.upper()} file :\n {os.path.basename(self.file_path)}\n and is being copied into a separate .XLSX formatted file\n', 'NOTICE: Conversion to Spreadsheet Format', wx.OK | wx.ICON_INFORMATION)
				rtn = self.xsv_to_xlsx(file_extension)
				logging.debug(f'UI: {file_extension.upper()}  conversion succeeded. New file is: ' + os.path.basename(self.file_path))
				wx.MessageBox(f"{file_extension.upper()} conversion to XLSX Completed\n File saved as {os.path.basename(self.file_path)}", "SUCCESS XLSX Conversion Complete", wx.OK | wx.ICON_INFORMATION)

			
			self.spreadsheet = pd.ExcelFile(self.file_path)
			self.tab_choice.Set(self.spreadsheet.sheet_names)
			self.file_name=os.path.basename(self.file_path)
			self.file_dir_only=os.path.dirname(self.file_path)
			#logging.debug('got here you')
			for radio_button in self.radio_buttons:
				radio_button.Enable()
				radio_button.SetValue(False)  
		else: # if no file selected, disable the radio buttons that inform the mapping format needed to be used
			for radio_button in self.radio_buttons:
				radio_button.Disable()
		
	def on_tab_select(self, event):
		logging.debug('UI: Getting the tab name that was selected')
		self.selected_tab = self.tab_choice.GetString(self.tab_choice.GetSelection())
		#Panda method for getting count
		if self.spreadsheet and self.selected_tab:
			df = pd.read_excel(self.spreadsheet, sheet_name=self.selected_tab, engine='openpyxl')
			df = df.dropna(how='all')  # Drop rows where all elements are NaN
			#last_row_number = df.index[-1] # Get the row number of the last row with data
			#last_row_number=self.getCount()
			#self.row_count_label.SetLabel(f"Row Count: {row_count}")
		
			#Use openpyxl for looping - pandas iterations gives odd numbers
			ref_workbook = openpyxl.load_workbook(self.file_path, read_only=True)
			source_sheet = ref_workbook[self.selected_tab]   #self.file_Max_Row = len(source_sheet['A'])
			last_row_number = source_sheet.max_row
			logging.debug("PD row count = " + str(last_row_number))
			
			logging.debug('UI: Getting the Header row number if possible. Assumes all file have a column header called "MRI Song ID" or "MLC Song Code"')
			rowFound = False
			totalFound = False
			mlc=False
			msgText=""			
			for row in source_sheet.rows:
				for cell in row:
					if cell.value == 'MRI Song ID' or cell.value == 'MLC Song Code' or cell.value == 'Distribution Identifier' or cell.value == 'Client Code':
						self.header_row.SetValue(str(cell.row))
						#logging.debug(self.header_row.GetValue())
						rowFound = True
						if cell.value == 'Distribution Identifier' or cell.value=='Client Code': 
							mlc=True 
						else: 
							mlc=False
						break
						
			if not rowFound: 
				msgText = 'Could not find the row with the column headers.'
				# wx.MessageBox(msgText, 'Source File Issue Warning', wx.OK | wx.ICON_WARNING)
				self.header_row.SetValue(str(0))
						
			logging.debug('UI: Getting the number of rows if possble; this would be the last row of data to be transformed. Assumes every file has a "Total" row')
			if mlc is True: 
				self.row_count_text.SetValue(str(last_row_number))
				totalFound = True
			else: 
				for row in source_sheet.rows:
					for cell in row:
						if cell.value == 'Total':
							self.row_count_text.SetValue(str(cell.row-1))
							totalFound = True
							break
						
			if not totalFound:
				msgText += '\n Could not determine the last good data row of the spreadsheet. \n Please verify estimated row count field value manually, if provided, or select a different tab'
				self.row_count_text.SetValue(str(last_row_number))
				
			if not rowFound or not totalFound:
				wx.MessageBox(msgText, 'Source File Issue Warning', wx.OK | wx.ICON_WARNING)
				msgText = msgText.strip('\n')
				logging.warning(f'UI: {msgText}')

	def on_radio_button(self, event):
		foundIncStmtType = False
		# Get the label of the selected radio button
		selected_radio = event.GetEventObject()
		self.radiolabel = selected_radio.GetLabel()
		
		incFileTypes = self.inc_stmt_search()
		
		if isinstance(incFileTypes, tuple):
			foundIncStmtType = any(incFileType.upper() in self.file_path.upper() for incFileType in incFileTypes)
		else:
			foundIncStmtType = self.radiolabel.upper() in self.file_path.upper()
		
# 		if self.radiolabel.upper() not in self.file_path.upper():
		if foundIncStmtType is False:
			rtn = wx.MessageBox('Please confirm your income statement selection,\n\n'  + self.radiolabel + '\n\n is correct for your file selection: \n\n' + self.file_name + '\n', 'Income Statement Type Selection WARNING', wx.OK | wx.ICON_WARNING)
			logging.warning(f'UI: Warning triggered regarding an unmatched Income Statement Type radio button select {self.radiolabel}')
			
	def inc_stmt_search(self):
		if self.radiolabel == "Amazon Music":
			return ("Amazon Music", "unlimited")
		elif self.radiolabel == "Amazon Prime":
			return ("Amazon Prime", "Prime_")
		elif self.radiolabel ==  "Amazon Prime Perf":
			return ("Amazon Prime Perf", "PrimePerf", "Perf")
		else:
			return self.radiolabel
		
	def perform_validations(self):
		logging.debug('UI: Performing validations...')
		msgText=''
		#logging.debug('path= ' + self.file_path)
		if not self.file_path: msgText +=' - You must select a spreadsheet to import \n'
		if not self.selected_tab: msgText +=' - A tab/sheet name must be selected \n'	
		if not self.radiolabel: msgText +=' - You must select an Income Statement Type \n'
		#if self.label == 'none' and self.MLC_IncomeStatement == 'none': msgText +=' - You must select an Income Statement Type \n'
		#if self.MRI_incomeStatement != 'none' and self.MLC_IncomeStatement != 'none': msgText +=' - You cannot select both MLC and MRI Income Statement Types \n'
		
		if len(msgText) > 1:
			wx.MessageBox(msgText, 'Validation Issues - Formatting Failure', wx.OK | wx.ICON_ERROR)
			return "failed"	

	def on_save(self, event):
		sam = self.perform_validations()
		if sam == 'failed':
			return
			
		if self.file_dir_only:
			save_df = pd.DataFrame({
				"File Path": [self.file_path], 
				"Directory": [self.file_dir_only], 
				"File Name": [self.file_name],
				"Tab": [self.selected_tab], 
				"HeaderRow": [self.header_row.GetValue()], 
				"LastDataRow": [self.row_count_text.GetValue()], 
				"Income Statement": [self.radiolabel], 
				"MLC": ["none"]
			})
			save_df.to_excel(self.file_dir_only + '/user_data.xlsx', index=False)
			wx.MessageBox("Details saved successfully!", "Info", wx.OK | wx.ICON_INFORMATION)
		
		# Run the Formatter code which will read from the newly created spreadsheet for the config info needed
		logging.debug("UI: dir the user chose where the data files are and where the output files will be written: " + self.file_dir_only)
		#logging.debug(os.getcwd())
		
		
		
		if getattr(sys, 'frozen', False):
			script_dir = os.path.dirname(sys.executable)
			logging.debug('UI: executable dir path chosen ' + script_dir)
		else:
			script_dir = os.path.dirname(os.path.abspath(__file__))
			logging.debug('UI: abspath or path given a file to that file ' + script_dir)
			
		# Change the working directory to the script's directory
		os.chdir(script_dir)
			
		# Now you can use relative paths
		#exec_file_path = os.path.join(script_dir, '/MRI_to_RyteBox_Formatter_082224.py')
		exec_file_path = script_dir + "\\MRI_to_RyteBox_Formatter_082224.py"
		
		logging.debug(f'UI: Path being passed to the FM call is: {exec_file_path}')
		logging.debug(f'UI: argument being passed to Formst is {self.file_dir_only}')
		
		try:
			logging.debug('UI: attempting to call the Format py now...')
				#bob = os.system(f'"{exec_file_path}" {self.file_dir_only}')
				#bob = os.system(f'"python" ".\\MRI_to_RyteBox_Formatter_082224.py" "{self.file_dir_only}"')
			bob = MRI_to_RyteBox_Formatter_082224.main(Path(self.file_dir_only)) 
			logging.debug(f'UI: Formatter Response or error code str({bob})')
			if bob is None:
				wx.MessageBox(f"Income Statement File: {self.file_name} Formatted Successfully to RyteBox Standard Format!", "Success", wx.OK | wx.ICON_INFORMATION)
			else:
				wx.MessageBox(f"Failed to process file: {self.file_name} due to error {bob}", "ERROR", wx.OK | wx.ICON_ERROR)
				logging.error(f'Failed to process file: {self.file_name} due to error {bob}')
			logging.debug('UI: finished call to the Format PY')
		except subprocess.CalledProcessError as e:
			logging.error("UI: Error:", e.stderr)
			logging.error("UI: Return code:", e.returncode)
			wx.MessageBox(f"Failed to process file: {self.file_name} due to {e}", "ERROR", wx.OK | wx.ICON_ERROR)
		return
		
	def xsv_to_xlsx(self, file_extension):
		#Convert TSV or CSV to XLSX while preserving cell text-format to preserve leading zeros and prevent excel sci notation on floats 
		xsvFile = self.file_path
		if file_extension.upper() == '.CSV': 
			# Read the .csv file
			df = pd.read_csv(xsvFile, dtype=str)
		else: # Rea the .tsv file
			logging.debug('UI: Performing TSV file translation to protected XLSX')
			df = pd.read_csv(xsvFile, delimiter='\t', dtype=str)
		
		xlsx_path=self.file_path.replace(file_extension, '.xlsx')
		
		# Create a new Excel writer object using XlsxWriter as the engine
		writer = pd.ExcelWriter(xlsx_path, engine='xlsxwriter')
	
		# Convert the dataframe to an XlsxWriter Excel object
		df.to_excel(writer, index=False, sheet_name='Sheet1')
		# Get the XlsxWriter workbook and worksheet objects
		workbook  = writer.book
		worksheet = writer.sheets['Sheet1']
		# Define a format to preserve leading zeros
		text_format = workbook.add_format({'num_format': '@'})
		# Apply the text format to all columns
		for col_num, value in enumerate(df.columns.values):
			worksheet.set_column(col_num, col_num, None, text_format)
		
		writer.save() # Save the workbook
		self.file_picker.SetPath(xlsx_path)
		self.file_path = xlsx_path
	
	def getCount(self):
		workbook = load_workbook(self.spreadsheet)
		sheet = workbook(self.selected_tab)
		return sheet.max_row
		
	def OnClose(self, event):     # added to address the issue of WX firing events when the user attempts to close the form/exit the app
		# Perform any cleanup here
		self.Destroy()
		event.Skip()
		

if __name__ == "__main__":
	app = wx.App(False)
	frame = SpreadsheetApp(None, "Select Income Statement To Format")
	try:
		app.MainLoop()
	except Exception as e:
		logging.error("Main: Exception occurred", exc_info=True)