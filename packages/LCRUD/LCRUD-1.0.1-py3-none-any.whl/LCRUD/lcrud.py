# lcrud.py
import openpyxl

# Função para carregar o arquivo Excel
def Load(path):
    global wb
    wb = openpyxl.load_workbook(path)
    global sheet
    sheet = wb.active
    global file_path
    file_path = path

# Função para listar os dados
def list(row=None, column=None):
    if row is None and column is None:
        # Se nenhum argumento for passado, retorna todos os dados da planilha
        rows = sheet.iter_rows(values_only=True)
        all_data = [row_data for row_data in rows]
        return all_data
    elif row is not None and column is None:
        # Se apenas o número da linha for passado, retorna todos os dados dessa linha
        row_data = sheet.iter_rows(min_row=row, max_row=row, values_only=True)
        return [data for data in row_data][0]
    elif row is not None and column is not None:
        # Se o número da linha e coluna forem passados, retorna o valor da célula específica
        cell_value = sheet.cell(row=row, column=column).value
        return cell_value

# Função para atualizar
def UpdateFile(row, column, value):
    sheet.cell(row=row, column=column).value = value
    wb.save(file_path)

# Função para deletar
def delete(row, column):
    sheet.cell(row=row, column=column).value = None
    wb.save(file_path)
