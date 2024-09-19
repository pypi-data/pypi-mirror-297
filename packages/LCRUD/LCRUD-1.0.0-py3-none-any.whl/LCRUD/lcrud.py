# LCRUD.py

import openpyxl

# Função para carregar o arquivo Excel
def Load(path):
    global wb
    wb = openpyxl.load_workbook(path)
    global sheet
    sheet = wb.active
    global file_path
    file_path = path

# Função para listar os dados
def list(row=None, column=None):
    if row is None and column is None:
        rows = sheet.iter_rows(values_only=True)
        for row_data in rows:
            print(row_data)
    elif row is not None and column is None:
        row_data = sheet.iter_rows(min_row=row, max_row=row, values_only=True)
        for data in row_data:
            print(data)
    elif row is not None and column is not None:
        cell_value = sheet.cell(row=row, column=column).value
        print(cell_value)

# Função para atualizar
def UpdateFile(row, column, value):
    sheet.cell(row=row, column=column).value = value
    wb.save(file_path)

# Função para deletar
def delete(row, column):
    sheet.cell(row=row, column=column).value = None
    wb.save(file_path)
