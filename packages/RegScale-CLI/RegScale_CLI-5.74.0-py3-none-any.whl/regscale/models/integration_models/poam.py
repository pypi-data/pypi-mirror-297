#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" A class to import Fedramp V4 and V5 POAMS """
import re
from collections import Counter
from pathlib import Path
from typing import List, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
from rich.console import Console

from regscale.core.app.application import Application
from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import create_progress_object
from regscale.core.utils.date import date_str, datetime_str
from regscale.integrations.integration.issue import IntegrationIssue
from regscale.models import IssueSeverity
from regscale.models.regscale_models import Issue


class POAM(IntegrationIssue):
    """
    Custom Integration issue class
    """

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        logger = create_logger()
        self.logger = logger
        file_path: Union[str, None] = kwargs.get("file_path", None)
        if not file_path:
            raise ValueError("File path is required")
        console = Console()
        self.file_path = Path(file_path)
        app = Application()
        self.app = app
        self.module = kwargs.get("module", "securityplans")
        self.module_id = kwargs.get("module_id", 0)
        self.console = console
        self.poam_data: List[Issue] = []
        data = self.import_poam()
        self.data = data
        self.create_or_update_issues(issues=self.poam_data, parent_id=self.module_id, parent_module=self.module)
        self.logger.info("Finished importing POAMs..")

    def pull(self):
        """
        Pull inventory from an Integration platform into RegScale
        """
        # Implement the pull method here
        pass

    def file_type(self):
        """
        A method to return the file type
        """
        file_type = None
        if self.file_path:
            file_type = self.file_path.suffix
        return file_type

    @staticmethod
    def get_index_from_column_name(column_name: str) -> int:
        """
        A method to get the index from a column name

        :param str column_name: A column name
        :return: The index of the column
        :rtype: int
        """
        return column_index_from_string(column_name) - 1

    def get_row_val(self, row: tuple, column_name: str) -> str:
        """
        Get the value from the row

        :param tuple row: The row
        :param str column_name: The column name
        :return: The value or None
        :rtype: str
        """
        return row[self.get_index_from_column_name(column_name)]

    def get_basis_for_adjustment(self, row: tuple) -> Optional[str]:
        """
        Get the basis for adjustment

        :param tuple row: The row
        :return: The basis for adjustment or None if adjusted risk rating is the same as risk rating
        :rtype: Optional[str]
        """
        basis_for_adjustment = self.empty(row[23])
        risk_rating = self.get_row_val(row, "S")
        adjusted_risk_rating = self.get_row_val(row, "T")
        if (adjusted_risk_rating != risk_rating) and not basis_for_adjustment:
            return "POAM Import"
        if adjusted_risk_rating == risk_rating:
            return None
        return basis_for_adjustment

    def get_supporting_documents(self, row: tuple) -> str:
        """
        Get the supporting documents

        :param tuple row: The row
        :return: The supporting documents
        :rtype: str
        """
        return f"\nSupporting Documents: {self.get_row_val(row, 'Y')}" if self.empty(self.get_row_val(row, "Y")) else ""

    def gen_issue_from_row(self, row: tuple, status: str, category: str, index: int, sheet: str) -> Optional[Issue]:
        """
        A method to generate an issue from a row

        :param tuple row: A row
        :param str status: The status of the issue
        :param str category: The category of the issue
        :param int index: The index of the row
        :param str sheet: The sheet name
        :return: An issue or None
        :rtype: Optional[Issue]
        """
        false_positive = self.set_false_positive(row)

        basis_for_adjustment = self.get_basis_for_adjustment(row)
        original_risk_rating = self.empty(self.get_row_val(row, "S"))
        adjusted_risk_rating = self.get_row_val(row, "T")
        risk_adjustment = self.set_risk_adjustment(row)
        if adjusted_risk_rating == "N/A":
            adjusted_risk_rating = original_risk_rating or "N/A"

        cve = self.empty(self.get_row_val(row, "AD"))
        if cve and not cve.upper().startswith("CVE"):
            cve = ""
        poam_id = self.get_row_val(row, "A")
        if not poam_id or not poam_id.upper().startswith("V"):
            return
        plugin_name = self.get_row_val(row, "E")
        plugin_id = self.get_row_val(row, "F")
        date_created = date_str(self.get_row_val(row, "K"))
        date_last_updated = datetime_str(self.get_row_val(row, "O"))
        due_date = self.get_row_val(row, "L")
        if due_date == "#REF!":
            due_date = None
        weakness_name = self.get_row_val(row, "C")
        if not weakness_name:
            self.logger.warning("Title is required on row %i, sheet %s. Unable to import", index, sheet)
            return None
        supporting_documents = self.get_supporting_documents(row)
        severity_level = IssueSeverity.NotAssigned
        if category.title() in IssueSeverity.__members__:
            severity_level = getattr(IssueSeverity, category.title())

        issue = Issue(
            otherIdentifier=self.empty(poam_id),
            dateCreated=date_created,
            dateLastUpdated=date_last_updated,
            title=weakness_name[:255],
            description=f"{self.get_row_val(row, 'D')}{supporting_documents}",
            status=status,
            severityLevel=severity_level,
            assetIdentifier=self.get_row_val(row, "G"),
            isPoam=True,
            issueOwnerId=self.app.config["userId"],
            securityPlanId=self.module_id if self.module == "securityplans" else 0,
            cve=cve,
            sourceReport=plugin_name,
            pluginId=str(plugin_id),
            autoApproved="No",
            dueDate=date_str(due_date),
            parentId=self.module_id,  # type: ignore
            parentModule=self.module,  # type: ignore
            basisForAdjustment=basis_for_adjustment,
            dateCompleted=date_str(date_last_updated) if status == "Closed" else None,
            manualDetectionSource=plugin_name,
            manualDetectionId=str(plugin_id),
            changes=self.get_row_val(row, "N"),
            poamComments=self.empty(self.get_row_val(row, "Z")),
            deviationRationale=self.empty(self.get_row_val(row, "X")),
            remediationDescription=self.empty(self.get_row_val(row, "J")),
            vendorDependency=self.empty(self.get_row_val(row, "P")),
            vendorLastUpdate=self.empty(date_str(self.get_row_val(row, "Q"))),
            vendorName=self.empty(self.get_row_val(row, "R")),
            adjustedRiskRating=adjusted_risk_rating,
            originalRiskRating=original_risk_rating,
            falsePositive=false_positive,
            identification="Vulnerability Assessment",
            operationalRequirement=self.set_operational_requirement(row),
            dateFirstDetected=date_str(self.get_row_val(row, "K")),
            riskAdjustment=risk_adjustment,
        )
        return issue

    def import_poam(self) -> Workbook:
        """
        A method to import the POAM data

        :return: The workbook
        :rtype: Workbook
        """
        workbook = load_workbook(filename=self.file_path, data_only=True, read_only=True)
        sheets = workbook.sheetnames
        pattern = "POA&M Items"
        status = "Open"

        poam_sheets = [item for item in sheets if re.search(pattern, item)]
        with create_progress_object() as progress:
            parsing_progress = progress.add_task("[#f8b737]Parsing data from workbook...", total=len(poam_sheets))
            for sheet in poam_sheets:
                ws = workbook[sheet]
                parsing_poams = progress.add_task(
                    f"[#ef5d23]Parsing '{sheet}' sheet for POAMs...", total=ws.max_row - 7
                )
                category = ws["C3"].value
                min_row = 0
                for index, row in enumerate(ws.iter_rows(min_row=min_row, max_row=ws.max_row, values_only=True)):
                    if "closed" in sheet.lower():
                        status = "Closed"
                    try:
                        issue = self.gen_issue_from_row(
                            row=row, status=status, category=category, index=index + min_row, sheet=sheet
                        )
                        if issue:
                            self.poam_data.append(issue)
                    except TypeError:
                        continue
                    progress.update(parsing_poams, advance=1)
                progress.update(parsing_progress, advance=1)
        # Count issues by status
        self.count_issues_by_status()
        return workbook

    def count_issues_by_status(self):
        """
        A method to count the issues and log the counts.
        """
        status_list = [issue.status for issue in self.poam_data if issue]
        status_counts = Counter(status_list)
        self.logger.info(
            "Found %i issues in the POAM Workbook, %i Open and %i Closed.",
            len(self.poam_data),
            status_counts["Open"],
            status_counts["Closed"],
        )

    def empty(self, string: str) -> Union[str, None]:
        """
        A method to empty the data

        :param str string: A string
        :return: None if the string is 'None' or the input is not a string
        :rtype: Union[str, None]
        """
        if not isinstance(string, str):
            return None

        if string.lower() in ["none", "n/a"]:
            return None

        return string

    def set_false_positive(self, row: tuple) -> str:
        """
        Set the false positive value

        :param tuple row: The row
        :return: The false positive value
        :rtype: str
        """
        # Map lowercased values to their corresponding responses
        value_map = {"yes": "Yes", "no": "No", "pending": "Pending Review"}

        # Get the value from the row and convert it to lowercase
        if row_value := self.get_row_val(row, "V"):
            row_value = row_value.lower()

        # Get the corresponding response from the map, default to 'No' if not found
        return value_map.get(row_value, "No")

    def set_operational_requirement(self, row: tuple) -> str:
        """
        Set the operational requirement value

        :param tuple row: The row
        :return: The operational requirement value
        :rtype: str
        """
        # Map lowercased values to their corresponding responses
        value_map = {"yes": "Yes", "no": "No", "pending": "Pending"}

        # Get the value from the row and convert it to lowercase
        if row_value := self.get_row_val(row, "W"):
            row_value = row_value.lower()

        # Get the corresponding response from the map, default to No if not found
        return value_map.get(row_value, "No")

    def set_risk_adjustment(self, row: tuple) -> str:
        """
        Set the risk adjustment value

        :param tuple row: The row
        :return: The Risk adjustment string
        :rtype: str
        """
        value_map = {"yes": "Yes", "no": "No", "pending": "Pending"}
        if row_value := self.get_row_val(row, "U"):
            row_value = row_value.lower()
        return value_map.get(row_value, "No")
